import openpyxl

path = r'C:\Users\HP\Desktop\Finalproject\Student input data.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active
print('SHEETS', wb.sheetnames)
print('ROWS', ws.max_row, 'COLS', ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=min(12, ws.max_row), values_only=True):
    print(row)
