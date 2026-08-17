import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

from app.database import SessionLocal, init_db
from app.models.geography import School
from app.models.student import Student
from seed_geography import seed_geography


WORKBOOK_PATH = Path(r"C:\Users\HP\Desktop\Finalproject\Student input data.xlsx")
SHEET_NAMES = ["teachers_remarks", "counsellor_remarks"]


def normalize_student_id(raw_value, fallback_index):
    raw = str(raw_value or "").strip()
    if not raw:
        return f"SK{fallback_index:04d}"

    if raw.upper().startswith("SMK"):
        return "SK" + raw[3:].strip()
    if raw.upper().startswith("SK"):
        return raw.upper()

    digits = re.sub(r"\D", "", raw)
    if not digits:
        return f"SK{fallback_index:04d}"
    return f"SK{digits}"


def parse_date_from_ic(raw_value):
    digits = re.sub(r"\D", "", str(raw_value or ""))
    if len(digits) < 6:
        return date(2000, 1, 1)

    try:
        dd = int(digits[0:2])
        mm = int(digits[2:4])
        yy = int(digits[4:6])
        if yy < 30:
            yy += 2000
        else:
            yy += 1900
        return date(yy, mm, dd)
    except ValueError:
        return date(2000, 1, 1)


def parse_gender_from_ic(raw_value):
    digits = re.sub(r"\D", "", str(raw_value or ""))
    if len(digits) < 9:
        return "unknown"
    return "Male" if int(digits[8]) % 2 == 1 else "Female"


def load_students():
    init_db()
    seed_geography()

    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    records = []

    for sheet_name in SHEET_NAMES:
        sheet = workbook[sheet_name]
        header = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(header, row))
            student_ic = data.get("Student IC")
            if not student_ic:
                continue
            if str(student_ic).strip() in {str(r["student_ic"]) for r in records}:
                continue

            name = data.get("Student Name") or "Unknown"
            records.append({
                "student_ic": str(student_ic).strip(),
                "student_id": normalize_student_id(student_ic, len(records) + 1),
                "full_name": str(name).strip(),
                "date_of_birth": parse_date_from_ic(student_ic),
                "gender": parse_gender_from_ic(student_ic),
                "class_name": "Unknown",
                "school_year": datetime.now().year,
                "socioeconomic_status": "unknown",
                "guardian_name": None,
                "guardian_contact": None,
            })

    db = SessionLocal()
    try:
        school = db.query(School).order_by(School.id).first()
        if not school:
            raise RuntimeError("No school found; geography seed did not create a school.")

        inserted = 0
        for record in records:
            existing = db.query(Student).filter(Student.student_id == record["student_id"]).first()
            if existing:
                continue
            student = Student(
                student_id=record["student_id"],
                full_name=record["full_name"],
                date_of_birth=record["date_of_birth"],
                gender=record["gender"],
                class_name=record["class_name"],
                school_year=record["school_year"],
                school_id=school.id,
                socioeconomic_status=record["socioeconomic_status"],
                guardian_name=record["guardian_name"],
                guardian_contact=record["guardian_contact"],
            )
            db.add(student)
            inserted += 1

        db.commit()
        print(f"Imported {inserted} students into the database.")
    finally:
        db.close()


if __name__ == "__main__":
    load_students()
